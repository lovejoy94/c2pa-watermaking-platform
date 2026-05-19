





# ================================================================
# FICHIER  : services/watermark_document.py
# ROLE     : Microservice watermark documents
# PORT     : 5006
# LANCER   : python -m uvicorn services.watermark_document:app --port 5006 --reload
# ================================================================

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import os
import json
import uuid
import shutil
import hashlib
from datetime import datetime

app = FastAPI(
    title="Watermark Document Service",
    description="Détecte et applique le watermark sur les documents",
    version="1.0.0"
)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "outputs"

# ================================================================
# UTILITAIRES
# ================================================================

def watermark_any_file(input_path: str, output_path: str) -> str:

    manifest = {
        "manifest_id": "c2pa-" + hashlib.md5(str(datetime.now()).encode()).hexdigest()[:8],
        "creator": "C2PA-Service-Groupe7",
        "created_date": datetime.now().isoformat(),
        "media_type": "document"
    }

    with open(input_path, "rb") as f:
        manifest["file_hash"] = hashlib.sha256(f.read()).hexdigest()

    ext = os.path.splitext(input_path)[1].lower()

    try:

        if ext == ".pdf":

            try:
                from PyPDF2 import PdfReader, PdfWriter

                reader = PdfReader(input_path)

                writer = PdfWriter()

                for page in reader.pages:
                    writer.add_page(page)

                writer.add_metadata({
                    "/Title": f"C2PA-{manifest['manifest_id'][:8]}"
                })

                with open(output_path, "wb") as f:
                    writer.write(f)

            except Exception:
                shutil.copy2(input_path, output_path)

        elif ext == ".docx":

            try:
                from docx import Document

                doc = Document(input_path)

                doc.core_properties.title = (
                    f"C2PA-{manifest['manifest_id'][:8]}"
                )

                doc.save(output_path)

            except Exception:
                shutil.copy2(input_path, output_path)

        elif ext == ".xlsx":

            try:
                from openpyxl import load_workbook

                wb = load_workbook(input_path)

                wb.properties.title = (
                    f"C2PA-{manifest['manifest_id'][:8]}"
                )

                wb.save(output_path)

            except Exception:
                shutil.copy2(input_path, output_path)

        elif ext == ".pptx":

            try:
                from pptx import Presentation

                prs = Presentation(input_path)

                prs.core_properties.title = (
                    f"C2PA-{manifest['manifest_id'][:8]}"
                )

                prs.save(output_path)

            except Exception:
                shutil.copy2(input_path, output_path)

        else:

            shutil.copy2(input_path, output_path)

    except Exception:

        shutil.copy2(input_path, output_path)

    # ============================================================
    # MANIFEST JSON
    # ============================================================

    with open(output_path + ".c2pa.json",
              "w",
              encoding="utf-8") as f:

        json.dump(
            manifest,
            f,
            indent=2,
            ensure_ascii=False
        )

    return output_path

def verify_document_watermark(filepath: str,
                              filename: str) -> dict:

    result = {

        "success": True,

        "fichier": filename,

        "type_media": "document",

        "watermark_found": False,

        "confidence": 0,

        "details": [],

        "error": None
    }

    try:

        json_path = filepath + ".c2pa.json"

        # ========================================================
        # JSON C2PA
        # ========================================================

        if os.path.exists(json_path):

            result["watermark_found"] = True

            result["confidence"] = 100

            result["details"].append(
                "Manifeste C2PA document détecté"
            )

            return result

        ext = os.path.splitext(filepath)[1].lower()

        # ========================================================
        # PDF
        # ========================================================

        if ext == ".pdf":

            try:
                from PyPDF2 import PdfReader

                reader = PdfReader(filepath)

                meta = reader.metadata

                if (
                    meta
                    and "/Title" in meta
                    and "C2PA" in str(meta["/Title"])
                ):

                    result["watermark_found"] = True

                    result["confidence"] = 90

                    result["details"].append(
                        "Watermark C2PA détecté dans les métadonnées PDF"
                    )

            except Exception as e:

                result["details"].append(
                    f"Erreur lecture PDF : {str(e)}"
                )

        # ========================================================
        # DOCX
        # ========================================================

        elif ext == ".docx":

            try:
                from docx import Document

                doc = Document(filepath)

                if "C2PA" in str(doc.core_properties.title):

                    result["watermark_found"] = True

                    result["confidence"] = 90

                    result["details"].append(
                        "Watermark C2PA détecté dans les métadonnées DOCX"
                    )

            except Exception as e:

                result["details"].append(
                    f"Erreur lecture DOCX : {str(e)}"
                )

        # ========================================================
        # XLSX
        # ========================================================

        elif ext == ".xlsx":

            try:
                from openpyxl import load_workbook

                wb = load_workbook(filepath)

                if "C2PA" in str(wb.properties.title):

                    result["watermark_found"] = True

                    result["confidence"] = 90

                    result["details"].append(
                        "Watermark C2PA détecté dans les métadonnées XLSX"
                    )

            except Exception as e:

                result["details"].append(
                    f"Erreur lecture XLSX : {str(e)}"
                )

        # ========================================================
        # PPTX
        # ========================================================

        elif ext == ".pptx":

            try:
                from pptx import Presentation

                prs = Presentation(filepath)

                if "C2PA" in str(prs.core_properties.title):

                    result["watermark_found"] = True

                    result["confidence"] = 90

                    result["details"].append(
                        "Watermark C2PA détecté dans les métadonnées PPTX"
                    )

            except Exception as e:

                result["details"].append(
                    f"Erreur lecture PPTX : {str(e)}"
                )

        # ========================================================
        # AUCUN WATERMARK
        # ========================================================

        if not result["watermark_found"]:

            result["details"].append(
                "Aucun watermark détecté"
            )

    except Exception as e:

        result["success"] = False

        result["error"] = str(e)

        result["details"].append(
            f"Erreur analyse : {str(e)}"
        )

    return result

# ================================================================
# ROUTES
# ================================================================

@app.post("/watermark-document/detect")
async def detect_document(media: UploadFile = File(...)):

    if not media.filename:

        raise HTTPException(
            status_code=400,
            detail="Aucun fichier envoyé"
        )

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    unique_name = f"{uuid.uuid4()}_{media.filename}"

    filepath = os.path.join(
        UPLOAD_FOLDER,
        unique_name
    )

    with open(filepath, "wb") as f:

        shutil.copyfileobj(
            media.file,
            f
        )

    try:

        result = verify_document_watermark(
            filepath,
            media.filename
        )

        return JSONResponse(result)

    finally:

        try:

            if os.path.exists(filepath):

                os.remove(filepath)

            if os.path.exists(filepath + ".c2pa.json"):

                os.remove(filepath + ".c2pa.json")

        except Exception as e:

            print(
                f"Erreur suppression fichier : {e}"
            )

@app.post("/watermark-document/add")
async def add_document(media: UploadFile = File(...)):

    if not media.filename:

        raise HTTPException(
            status_code=400,
            detail="Aucun fichier envoyé"
        )

    os.makedirs(UPLOAD_FOLDER, exist_ok=True)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    unique_name = f"{uuid.uuid4()}_{media.filename}"

    filepath = os.path.join(
        UPLOAD_FOLDER,
        unique_name
    )

    output_path = os.path.join(
        OUTPUT_FOLDER,
        "wm_" + media.filename
    )

    with open(filepath, "wb") as f:

        shutil.copyfileobj(
            media.file,
            f
        )

    try:

        watermark_any_file(
            filepath,
            output_path
        )

        return JSONResponse({

            "success": True,

            "fichier": media.filename,

            "type_media": "document",

            "output": output_path,

            "error": None
        })

    except Exception as e:

        return JSONResponse({

            "success": False,

            "fichier": media.filename,

            "type_media": "document",

            "output": None,

            "error": str(e)

        }, status_code=500)

    finally:

        try:

            if os.path.exists(filepath):

                os.remove(filepath)

        except Exception as e:

            print(
                f"Erreur suppression fichier : {e}"
            )

# ================================================================
# HEALTH
# ================================================================

@app.get("/health")
async def health():

    return {

        "service": "watermark_document",

        "status": "ok",

        "port": 5006
    }

# ================================================================
# RUN
# ================================================================

if __name__ == "__main__":

    import uvicorn

    uvicorn.run(
        "services.watermark_document:app",
        host="0.0.0.0",
        port=5006,
        reload=True
    )
